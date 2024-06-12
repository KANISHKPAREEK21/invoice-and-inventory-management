from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect

from utils.filehandler import handle_file_upload
from openpyxl import load_workbook
from pypdf import PdfReader 
    
from .forms import *
from .models import *
import pandas as pd

from django.contrib.auth.decorators import login_required

# Create your views here.
@login_required
def home(request):
    return redirect("create_invoice")

# @login_required
def getTotalIncome():
    allInvoice = Invoice.objects.filter(invoice_is_delete=0)
    totalIncome = 0
    for curr in allInvoice:
        totalIncome += curr.total
    return round(totalIncome,2)


@login_required
def base(request):
    total_product = Product.objects.filter(product_is_delete=0).count()


    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()
    context = {
        "total_product": total_product,

        # this is commented
        "total_customer": total_customer,
        # this is commented

        "total_invoice": total_invoice,
        "total_income": total_income,
    }

    return render(request, "invoice/base/base.html", context)


@login_required
def download_all(request):
    # Download all invoice to excel file
    # Download all product to excel file
    # Download all customer to excel file

    allInvoiceDetails = InvoiceDetail.objects.all()
    invoiceAndProduct = {
        "invoice_id": [],
        "invoice_date": [],
        "invoice_customer": [],
        "invoice_contact": [],
        "invoice_comments": [],
        "product_name": [],
        # "product_price": [],
        "product_unit": [],
        # "product_amount": [], 
        "invoice_total": [],

    }
    for curr in allInvoiceDetails:
        invoice_detail_id = curr.id
        amount = curr.amount
        invoice_id = curr.invoice_id
        # print(invoice_id)
        product_id = curr.product_id
        price = curr.price

        invoice = Invoice.objects.filter(id = invoice_id)
        for i in invoice:
        # print(invoice['contact'])
            date = i.date
            customer_id = i.customer_id
            customer_comments = i.comments
            customer_contact = i.contact

        customer = Customer.objects.filter(id = customer_id)
        for i in customer:
            customer_name = i.customer_name

        product = Product.objects.filter(id = product_id)
        for i in product:
            product_name = i.product_name
            product_unit = i.product_unit
        # invoice = Invoice.objects.get(id=curr.id)
        # product = Product.objects.get(id=curr.product_id)
        # product = Product.objects.filter(id=curr.product_id)
        invoiceAndProduct["invoice_id"].append(invoice_id)
        invoiceAndProduct["invoice_date"].append(date)
        invoiceAndProduct["invoice_customer"].append(customer_name)
        invoiceAndProduct["invoice_contact"].append(customer_contact)
        # # invoiceAndProduct["invoice_email"].append(invoice.email)
        invoiceAndProduct["invoice_comments"].append(customer_comments)
        invoiceAndProduct["product_name"].append(product_name)
        # invoiceAndProduct["product_price"].append(product.product_price)
        invoiceAndProduct["product_unit"].append(product_unit)
        # invoiceAndProduct["product_amount"].append(curr.amount)
        invoiceAndProduct["invoice_total"].append(price)

    df = pd.DataFrame(invoiceAndProduct)
    df.to_excel("static/excel/allInvoices.xlsx", index=False, engine='openpyxl')

    # Load the workbook and the worksheet
    wb = load_workbook("static/excel/allInvoices.xlsx")
    ws = wb.active

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the adjusted file
    wb.save("static/excel/allInvoices.xlsx")

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="allInvoices.xlsx"'
    with open("static/excel/allInvoices.xlsx", "rb") as f:
        response.write(f.read())
    return response

@login_required
def download_invoice_detail(request , pk):
    # Download all invoice to excel file
    # Download all product to excel file
    # Download all customer to excel file

    # allInvoiceDetails = InvoiceDetail.objects.all()
    # total_product = Product.objects.filter(product_is_delete=0).count()

    # total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    # total_income = getTotalIncome()
    print("---------------Id for invoice--------------",pk)

    invoice = Invoice.objects.get(id=pk)
    print("---------------Id for invoice--------------",invoice)

    invoice_detail = InvoiceDetail.objects.filter(invoice_id=invoice)
    # if request.method == "POST":
        # invoice_detail.delete()
        # invoice.delete()
        # return redirect("view_invoice")

    invoiceAndProduct = {
        "invoice_id": [],
        "invoice_date": [],
        "invoice_customer": [],
        "invoice_contact": [],
        "invoice_comments": [],
        "product_name": [],
        "product_price": [],
        "product_amount": [],
        "product_unit": [],
        "invoice_total": [],
    }
    # for curr in allInvoiceDetails:
        # invoice = Invoice.objects.get(id=pk)
    cnt = 0
    for i in invoice_detail:
        product_id = i.product_id
        unit = i.unit

        product = Product.objects.filter(id = product_id)
        # products += product[0].product_name + ', '
        invoiceAndProduct["product_name"].append(product[0].product_name)
        # prices += str(product[0].product_price) + ', '
        invoiceAndProduct["product_price"].append(i.price)
        invoiceAndProduct["product_amount"].append(i.amount)
        invoiceAndProduct["product_unit"].append(unit)

        # product = Product.objects.get(id= invoice_detail.all().values()[0]['product_id'])
        # print(product)
        if cnt == 0:
            invoiceAndProduct["invoice_id"].append(invoice)
            invoiceAndProduct["invoice_date"].append(invoice.date)
            invoiceAndProduct["invoice_customer"].append(invoice.customer)
            invoiceAndProduct["invoice_contact"].append(invoice.contact)
            # invoiceAndProduct["invoice_email"].append(invoice.email)
            invoiceAndProduct["invoice_comments"].append(invoice.comments)
            # invoiceAndProduct["product_name"].append(products)
            # invoiceAndProduct["product_unit"].append(unit)
            # invoiceAndProduct["product_price"].append(prices)
            # invoiceAndProduct["product_amount"].append(invoice_detail.values()[0]['product_id'])
            invoiceAndProduct["invoice_total"].append(invoice.total)
            cnt = 1
        else:
            invoiceAndProduct["invoice_id"].append("")
            invoiceAndProduct["invoice_date"].append("")
            invoiceAndProduct["invoice_customer"].append("")
            invoiceAndProduct["invoice_contact"].append("")
            # invoiceAndProduct["invoice_email"].append(invoice.email)
            invoiceAndProduct["invoice_comments"].append("")
            # invoiceAndProduct["product_name"].append(products)
            # invoiceAndProduct["product_price"].append(prices)
            # invoiceAndProduct["product_unit"].append("")
            # invoiceAndProduct["product_amount"].append(invoice_detail.values()[0]['product_id'])
            invoiceAndProduct["invoice_total"].append("")


    df = pd.DataFrame(invoiceAndProduct)
    print(df)
    df.to_excel("static/excel/invoice.xlsx", index=False, engine='openpyxl')

    # Load the workbook and the worksheet
    wb = load_workbook("static/excel/invoice.xlsx")
    ws = wb.active

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the adjusted file
    wb.save("static/excel/invoice.xlsx")

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="invoice.xlsx"'
    with open(f"static/excel/invoice.xlsx", "rb") as f:
        response.write(f.read())
    return response

@login_required
def customer_invoice_download(request , pk):
    customer_details = {
        "customer_id":[],
        "customer_name":[],
        "customer_contact":[],
        "invoice_id":[],
        "invoice_date":[],
        "invoice_total":[],
        "is_GST_applied":[],
        "product_id":[],
        "product_name":[],
        "product_price":[],
        "product_amount":[],
        "product_unit":[],
        "customer_total":[]
    }
    customer = Customer.objects.filter(id = pk)
    customer_id = customer[0].id
    customer_name = customer[0].customer_name
    customer_contact = customer[0].customer_contact
    customer_total = customer[0].customer_amount
    # if cnt == 0:
    customer_details["customer_id"].append(customer_id)
    customer_details["customer_name"].append(customer_name)
    customer_details["customer_contact"].append(customer_contact)
    customer_details["customer_total"].append(customer_total)

    invoices = Invoice.objects.filter(customer_id = customer_id)
    for i in invoices:
        invoice_id = i.id
        invoice_date = i.date
        invoice_total = i.total
        gst = i.gst

        customer_details["invoice_id"].append(invoice_id)
        customer_details["invoice_date"].append(invoice_date)
        customer_details["invoice_total"].append(invoice_total)
        customer_details["is_GST_applied"].append(gst)

        invoice_detail = InvoiceDetail.objects.filter(invoice_id = invoice_id)
        cnt = 0
        for p in invoice_detail:
            product_id = p.product_id
            product_price = p.price
            product_amount = p.amount
            product_unit = p.unit

            customer_details["product_id"].append(product_id)
            products = Product.objects.filter(id = product_id)
            product_name = products[0].product_name

            customer_details["product_name"].append(product_name)
            customer_details["product_price"].append(product_price)
            customer_details["product_amount"].append(product_amount)
            customer_details["product_unit"].append(product_unit)

            if p != invoice_detail[len(invoice_detail) - 1]:
                customer_details["customer_id"].append("")
                customer_details["customer_name"].append("")
                customer_details["customer_contact"].append("")
                customer_details["customer_total"].append("")

                customer_details["invoice_id"].append("")
                customer_details["invoice_date"].append("")
                customer_details["invoice_total"].append("")
                customer_details["is_GST_applied"].append("")


            customer_details["customer_id"].append("")
            customer_details["customer_name"].append("")
            customer_details["customer_contact"].append("")
            customer_details["customer_total"].append("")

            customer_details["invoice_id"].append("")
            customer_details["invoice_date"].append("")
            customer_details["invoice_total"].append("")
            customer_details["is_GST_applied"].append("")

            customer_details["product_id"].append("")
            customer_details["product_name"].append("")
            customer_details["product_price"].append("")
            customer_details["product_amount"].append("")
            customer_details["product_unit"].append("")


        if i != invoices[len(invoices) - 1]:
            customer_details["customer_id"].append("")
            customer_details["customer_name"].append("")
            customer_details["customer_contact"].append("")
            customer_details["customer_total"].append("")

    print(customer_details)
    df = pd.DataFrame(customer_details)
    print(df)
    df.to_excel("static/excel/customer.xlsx", index=False, engine='openpyxl')

    # Load the workbook and the worksheet
    wb = load_workbook("static/excel/customer.xlsx")
    ws = wb.active

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the adjusted file
    wb.save("static/excel/customer.xlsx")

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="customer.xlsx"'
    with open(f"static/excel/customer.xlsx", "rb") as f:
        response.write(f.read())
    return response

# @login_required
# def delete_all_invoice(request):
#     # Delete all invoice
#     Invoice.objects.all().delete()
#     return redirect("view_invoice")


# def upload_product_from_excel(request):
#     # Upload excel file to static folder "excel"
#     # add all product to database
#     # save product to database
#     # redirect to view_product
#     excelForm = excelUploadForm(request.POST or None, request.FILES or None)
#     print("Reached HERE!")
#     if request.method == "POST":
#         print("Reached HERE2222!")

#         handle_file_upload(request.FILES["excel_file"])
#         excel_file = "static/excel/masterfile.xlsx"
#         df = pd.read_excel(excel_file)
#         Product.objects.all().delete()
#         for index, row in df.iterrows():
#             product = Product(
#                 product_name=row["product_name"],
#                 product_price=row["product_price"],
#                 # product_unit=row["product_unit"],
#             )
#             print(product)
#             product.save()
#         return redirect("view_product")
#     return render(request, "invoice/upload_products.html", {"excelForm": excelForm})

    # Product view


@login_required
def create_product(request):
    total_product = Product.objects.filter(product_is_delete=0).count()


    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    product = ProductForm()

    if request.method == "POST":
        product = ProductForm(request.POST)
        if product.is_valid():
            product.save()
            return redirect("view_product")

    context = {
        "total_product": total_product,

        # this is commented
        "total_customer": total_customer,
        # this is commented

        "total_invoice": total_invoice,
        "total_income": total_income,
        "product": product,
    }

    return render(request, "invoice/create_product.html", context)


@login_required
def view_product(request):
    total_product = Product.objects.filter(product_is_delete=0).count()

    
    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    product = Product.objects.filter(product_is_delete=False)
    print(product)
    context = {
        "total_product": total_product,
    
        # this is commented
        "total_customer": total_customer,
        # this is commented
    
        "total_invoice": total_invoice,
        "total_income": total_income,
        "product": product,
    }

    return render(request, "invoice/view_product.html", context)




        # this is commented
# Customer view
@login_required
def create_customer(request):
    total_product = Product.objects.filter(product_is_delete=0).count()

    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    customer = CustomerForm()

    if request.method == "POST":
        customer = CustomerForm(request.POST)
        if customer.is_valid():
            customer.save()
            return redirect("view_customer")

    context = {
        "total_product": total_product,
        "total_customer": total_customer,
        "total_invoice": total_invoice,
        "total_income" : total_income,
        "customer": customer,
    }

    return render(request, "invoice/create_customer.html", context)

        # this is commented

@login_required
def view_customer(request):
    total_product = Product.objects.filter(product_is_delete=0).count()

    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    customer = Customer.objects.filter(customer_is_delete=False)
    # customer = Customer.objects.all()
    for i in Customer.objects.all():
        income = 0 
        invoice = Invoice.objects.filter(customer=i.id, invoice_is_delete =  False)
        for j in invoice:
            income = income + j.total
        i.customer_amount = income
        i.save()

    context = {
        "total_product": total_product,
        "total_customer": total_customer,
        "total_invoice": total_invoice,
        "total_income" : total_income,
        "customer": customer,
    }

    return render(request, "invoice/view_customer.html", context)
        # this is commented


# Invoice view
@login_required
def create_invoice(request):
    total_product = Product.objects.filter(product_is_delete=0).count()

    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    form = InvoiceForm()
    product_form = ProductForm()
    product_formset = ProductFormSet()
    formset = InvoiceDetailFormSet()
    
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        formset = InvoiceDetailFormSet(request.POST)
        
        if form.is_valid():
            invoice = Invoice.objects.create(
                customer=form.cleaned_data.get("customer"),
                contact=form.cleaned_data.get("contact"),
                date=form.cleaned_data.get("date"),
                gst=form.cleaned_data.get("gst"),
                comments = form.cleaned_data.get("comments"),
                # price = form.cleaned_data.get("price")
            )
        if formset.is_valid():
            total = 0
            for form in formset:
                product = form.cleaned_data.get("product")
                amount = form.cleaned_data.get("amount")
                price = form.cleaned_data.get("price")
                if product and amount:
                    # Sum each row
                    sum = float(price) * float(amount)
                    unit = float(product.product_unit) * float(amount)
                    # Sum of total invoice
                    total += sum
                    InvoiceDetail(
                        invoice=invoice, product=product, amount=amount, price=price, unit= unit
                    ).save()
            if(invoice.gst == True):
                total += (0.18) * total 
        # this is commented

            print("costumer detail")
            print(invoice.customer)
            print(type(invoice.customer))

            # Pointing the customer
            points = 0
            if total > 1000:
                points += total / 1000
            
            # invoice.customer.customer_points = round(points)
            # # Save the points to Customer table
            # invoice.customer.save()
        # this is commented

            # Save the invoice
            # if(invoice.gst == True):
            #     total += (total / 100) * 18
            invoice.total = total
            invoice.save()
            
            for i in Customer.objects.all():
                income = 0 
                invoice = Invoice.objects.filter(customer=i.id)
                for j in invoice:
                    income = income + j.total
                i.customer_amount = income
                i.save()
            return redirect("view_invoice")

    context = {
        "total_product": total_product,

        # this is commented
        "total_customer": total_customer,
        # this is commented

        "total_invoice": total_invoice,
        "total_income": total_income,
        "form": form,
        "formset": formset,
        "product_form":product_form,
        "product_formset":product_formset,
    }

    return render(request, "invoice/create_invoice.html", context)


@login_required
def view_invoice(request):
    total_product = Product.objects.filter(product_is_delete=0).count()


    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    invoice = Invoice.objects.filter(invoice_is_delete=0)

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "invoice": invoice,
    }

    return render(request, "invoice/view_invoice.html", context)


# Detail view of invoices
@login_required
def view_invoice_detail(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()


        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    invoice = Invoice.objects.get(id=pk)
    invoice_detail = InvoiceDetail.objects.filter(invoice=invoice)

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        # 'invoice': invoice,
        "invoice_detail": invoice_detail,
    }

    print(context)

    return render(request, "invoice/view_invoice_detail.html", context)


# Delete invoice
@login_required
def delete_invoice(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    invoice = Invoice.objects.get(id=pk)
    invoice_detail = InvoiceDetail.objects.filter(invoice=invoice)
    if request.method == "POST":
        # invoice_detail.delete()
        # invoice.delete()
        invoice.invoice_is_delete = True
        invoice_detail.invoicedetail_is_delete = True   
        invoice.save()
        # invoice_detail.save()
        return redirect("view_invoice")

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "invoice": invoice,
        "invoice_detail": invoice_detail,
    }

    return render(request, "invoice/delete_invoice.html", context)


# Edit customer
@login_required
def edit_customer(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

    
    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented
    
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    customer = Customer.objects.get(id=pk)
    form = CustomerForm(instance=customer)

    if request.method == "POST":
        customer = CustomerForm(request.POST, instance=customer)
        if customer.is_valid():
            customer.save()
            return redirect("view_customer")

    context = {
        "total_product": total_product,
        "total_customer": total_customer,
        "total_invoice": total_invoice,
        "total_income": total_income,
        "customer": form,
    }

    return render(request, "invoice/create_customer.html", context)

        # this is commented
        
# Delete customer
@login_required
def delete_customer(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    customer = Customer.objects.get(id=pk)

    if request.method == "POST":
        customer.customer_is_delete = True
        customer.save()
        return redirect("view_customer")

    context = {
        "total_product": total_product,
        "total_customer": total_customer,
        "total_invoice": total_invoice,
        "total_income": total_income,
        "customer": customer,
    }

    return render(request, "invoice/delete_customer.html", context)
        # this is commented

# Edit product
@login_required
def edit_product(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    product = Product.objects.get(id=pk)
    form = ProductForm(instance=product)

    if request.method == "POST":
        # this is commented
        product = ProductForm(request.POST, instance=product)
        # this is commented

        product.save()
        return redirect("view_product")

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "product": form,
    }

    return render(request, "invoice/create_product.html", context)


# Delete product
@login_required
def delete_product(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    product = Product.objects.get(id=pk)

    if request.method == "POST":
        product.product_is_delete = True
        product.save()
        return redirect("view_product")

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "product": product,
    }

    return render(request, "invoice/delete_product.html", context)

def product_price(request, pk):
    if request.method == "POST":
        print("pk: ",pk)
        product = Product.objects.get(id=pk)
        price = product.product_price
        context = {
            "price": price
        }
        return JsonResponse(context)
    return HttpResponse('Something went wrong', status=500)

def customer_contact(request, pk):
    if request.method == "POST":
        print("pk: ",pk)
        customer = Customer.objects.get(id=pk)
        contact = customer.customer_contact
        context = {
            "contact": contact
        }
        return JsonResponse(context)
    return HttpResponse('Something went wrong', status=500)

@login_required
def add_expense(request):
    total_product = Product.objects.filter(product_is_delete=0).count()


    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    expense = ExpenseForm()

    if request.method == "POST":
        expense = ExpenseForm(request.POST)
        if expense.is_valid():
            expense.save()
            return redirect("view_expense")

    context = {
        "total_product": total_product,

        # this is commented
        "total_customer": total_customer,
        # this is commented

        "total_invoice": total_invoice,
        "total_income": total_income,
        "expense": expense,
    }

    return render(request, "invoice/add_expense.html", context)

@login_required
def view_expense(request):
    total = 0
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()
    expense = Expense.objects.filter(expense_is_active=True).all().values()
    for i in Expense.objects.filter(expense_is_active=True).values():
        total = total + i['expense_cost']
    print(total)
    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "expense": expense,
        "total":total,
    }
    return render(request, "invoice/view_expense.html", context)

@login_required
def delete_expense(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    expense = Expense.objects.get(id=pk)

    if request.method == "POST":
        expense.expense_is_active = False
        expense.save()
        return redirect("view_expense")

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "expense": expense,
    }

    return render(request, "invoice/delete_expense.html", context)


@login_required
def add_payment(request):
    total_product = Product.objects.filter(product_is_delete=0).count()

    # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
    # this is commented

    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    payment = PaymentForm()
    if request.method == "POST":
        payment = PaymentForm(request.POST)
        if payment.is_valid():
                    payment.save()
                    return redirect("view_payment")
    context = {
        "total_product": total_product,

        # this is commented
        "total_customer": total_customer,
        # this is commented

        "total_invoice": total_invoice,
        "total_income": total_income,
        "payment": payment,
    }

    return render(request, "invoice/add_payment.html", context)

@login_required
def view_payment(request):
    total = 0
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()
    payment = Payment.objects.filter(Payment_is_active=True).all().values()
    for i in Payment.objects.filter(Payment_is_active=True).values():
        total = total + i['Payment_cost']

    print(total)
    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "payment": payment,
        "total":total,
    }
    return render(request, "invoice/view_payment.html", context)

@login_required
def delete_payment(request, pk):
    total_product = Product.objects.filter(product_is_delete=0).count()

        # this is commented
    total_customer = Customer.objects.filter(customer_is_delete=0).count()
        # this is commented
    total_invoice = Invoice.objects.filter(invoice_is_delete=0).count()
    total_income = getTotalIncome()

    payment = Payment.objects.get(id=pk)

    if request.method == "POST":
        payment.Payment_is_active = False
        payment.save()
        return redirect("view_payment")

    context = {
        "total_product": total_product,
        # this is commented
        "total_customer": total_customer,
        # this is commented
        "total_invoice": total_invoice,
        "total_income": total_income,
        "payment": payment,
    }

    return render(request, "invoice/delete_payment.html", context)
